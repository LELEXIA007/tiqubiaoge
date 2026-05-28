import os
import re
import csv
from pathlib import Path


def extract_tables_from_txt(txt_file):
    """
    从记事本文件中提取表格信息
    :param txt_file: 记事本文件路径
    :return: 提取的表格列表，每个表格是一个二维列表
    """
    with open(txt_file, 'r', encoding='utf-8') as f:
        content = f.read()
    
    tables = []
    
    # 方法1：检测以制表符或空格分隔的表格
    lines = content.strip().split('\n')
    current_table = []
    
    for line in lines:
        line = line.strip()
        if not line:
            if current_table:
                tables.append(current_table)
                current_table = []
            continue
        
        # 检测是否为表格行（包含多个由空格或制表符分隔的列）
        # 使用正则表达式匹配由多个空格或制表符分隔的内容
        columns = re.split(r'\s{2,}|\t', line)
        # 过滤空列
        columns = [col.strip() for col in columns if col.strip()]
        
        if len(columns) > 1:  # 至少有两列才认为是表格
            current_table.append(columns)
        else:
            if current_table:
                tables.append(current_table)
                current_table = []
    
    # 处理最后一个表格
    if current_table:
        tables.append(current_table)
    
    # 方法2：检测Markdown格式的表格
    md_tables = re.findall(r'\|.*?\|\n\|.*?\|\n(?:\|.*?\|\n)*', content, re.DOTALL)
    for md_table in md_tables:
        table_lines = md_table.strip().split('\n')
        table = []
        for line in table_lines:
            if '---' in line:  # 跳过分隔线
                continue
            columns = [col.strip() for col in line.strip('|').split('|')]
            table.append(columns)
        if table:
            tables.append(table)
    
    return tables


def merge_similar_tables(tables):
    """
    合并同类表格
    :param tables: 表格列表
    :return: 合并后的表格列表
    """
    if not tables:
        return tables
    
    merged_tables = []
    # 按表格列数分组
    tables_by_columns = {}
    
    for table in tables:
        if not table:
            continue
        # 获取表格的列数
        num_columns = len(table[0]) if table else 0
        if num_columns not in tables_by_columns:
            tables_by_columns[num_columns] = []
        tables_by_columns[num_columns].append(table)
    
    # 合并相同列数的表格
    for num_columns, table_group in tables_by_columns.items():
        if len(table_group) == 1:
            merged_tables.append(table_group[0])
        else:
            # 合并表格，保留第一个表格的表头
            merged_table = [table_group[0][0]]  # 保留表头
            # 添加所有表格的数据行
            for table in table_group:
                merged_table.extend(table[1:])  # 跳过表头
            merged_tables.append(merged_table)
    
    return merged_tables

def save_tables_to_excel(tables, output_file):
    """
    将提取的表格保存到Excel文件中
    :param tables: 表格列表
    :param output_file: 输出Excel文件路径
    """
    # 合并同类表格
    merged_tables = merge_similar_tables(tables)
    print(f"合并后剩余 {len(merged_tables)} 个表格")
    
    # 尝试使用openpyxl库
    try:
        from openpyxl import Workbook
        
        wb = Workbook()
        
        for i, table in enumerate(merged_tables):
            if i == 0:
                ws = wb.active
                ws.title = f'Table {i+1}'
            else:
                ws = wb.create_sheet(title=f'Table {i+1}')
            
            for row_idx, row in enumerate(table, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(output_file)
        print(f"表格已成功保存到: {output_file}")
        
    except ImportError:
        # 安装openpyxl库
        print("正在安装openpyxl库...")
        import subprocess
        subprocess.check_call(["pip", "install", "openpyxl"])
        
        # 再次尝试保存
        from openpyxl import Workbook
        
        wb = Workbook()
        
        for i, table in enumerate(merged_tables):
            if i == 0:
                ws = wb.active
                ws.title = f'Table {i+1}'
            else:
                ws = wb.create_sheet(title=f'Table {i+1}')
            
            for row_idx, row in enumerate(table, 1):
                for col_idx, cell_value in enumerate(row, 1):
                    ws.cell(row=row_idx, column=col_idx, value=cell_value)
        
        wb.save(output_file)
        print(f"表格已成功保存到: {output_file}")


def main():
    print("=== 提取记事本文件中的表格信息 ===")
    
    # 指定要遍历的目录
    target_dir = r"C:\Users\86133\Desktop\提取表格\extracted_results"
    print(f"遍历目录: {target_dir}")
    
    # 检查目录是否存在
    if not os.path.exists(target_dir):
        print(f"错误：目录 {target_dir} 不存在")
        return
    
    # 收集所有.txt文件
    txt_files = []
    for root, dirs, files in os.walk(target_dir):
        for file in files:
            if file.endswith('.txt'):
                txt_files.append(os.path.join(root, file))
    
    if not txt_files:
        print("目录中未找到记事本文件")
        return
    
    print(f"找到 {len(txt_files)} 个记事本文件")
    
    # 提取所有文件中的表格
    all_tables = []
    for txt_file in txt_files:
        print(f"正在处理文件: {os.path.basename(txt_file)}")
        tables = extract_tables_from_txt(txt_file)
        if tables:
            print(f"  从该文件提取到 {len(tables)} 个表格")
            all_tables.extend(tables)
        else:
            print(f"  该文件中未检测到表格")
    
    if not all_tables:
        print("未检测到表格信息")
        return
    
    print(f"成功提取 {len(all_tables)} 个表格")
    
    # 生成输出文件路径
    output_file = os.path.join(target_dir, "all_tables.xlsx")
    
    # 保存表格
    print("正在保存表格...")
    save_tables_to_excel(all_tables, output_file)
    
    print("=== 提取完成 ===")


if __name__ == "__main__":
    main()